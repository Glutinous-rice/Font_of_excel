import openpyxl
from openpyxl.styles import Font
wb=openpyxl.Workbook()     #创建一个工作簿
ws=wb.active
fontx=Font(italic=True,name='Calibri',bold= True,size=24)   #把单元格属性赋值给fontx
ws['A1'].font=fontx                                         #把fontx传递给单元格A1
ws['A1']='Hello,World!'
wb.save('Styles.xlsx')     #注意一定不要在打开文件的时候再次储存文件，不然会报错


